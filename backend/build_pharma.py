"""
O'zbekiston dori-darmon reyestri (uzpharm-control.uz eksporti, Excel) ->
ixcham JSON. Runtime'da pharma.py shu JSON'ni yuklaydi (Excel/openpyxl runtime'da
kerak emas).

Ustunlar: 0 guvohnoma, 1 savdo nomi, 2 egasi, 3 ishlab chiqaruvchi, 4 davlat,
5 MNN(INN), 6 ATX, 7 farmakoguruh, 8 tur, 9 shakl+qadoq(doza), 10 modda, 11 retsept.
"""
import json
import re
import os
import openpyxl

_DIR = "/Users/macbook_uz/Downloads/N30-23.07.2026"
SRC = os.path.join(_DIR, "1. Отеч. и заруб. лек.ср.xlsx")
if not os.path.exists(SRC):
    SRC = "/Users/macbook_uz/Downloads/1. Отеч. и заруб. лек.ср.xlsx"
ANNUL = os.path.join(_DIR, "6. Аннул.лек.срва .xls")
OUT = os.path.join(os.path.dirname(__file__), "pharma_registry.json")


def parse_annulled():
    """#6: ro'yxatdan chiqarilgan (annullyatsiya) dorilar -> {norm_kalit: {name, atc}}."""
    out = {}
    if not os.path.exists(ANNUL):
        return out
    try:
        import xlrd
        ws = xlrd.open_workbook(ANNUL).sheet_by_index(0)
    except Exception:
        return out
    for r in range(2, ws.nrows):
        trade = str(ws.cell_value(r, 1))
        inn = str(ws.cell_value(r, 2)).strip()
        atc = (str(ws.cell_value(r, 7)).strip() if ws.ncols > 7 else "")
        cyr = re.split(r'[(\n]', trade)[0].strip()
        mlat = re.search(r'\(([^)\n]+)', trade)
        latin = mlat.group(1).strip() if mlat else ""
        display = (inn or latin or cyr).strip().title()
        entry = {"name": display, "atc": atc}
        for cand in [inn, (latin.split()[0] if latin else ""), cyr]:
            k = re.sub(r'[^a-z0-9]', '', ''.join(_CYR.get(ch, ch) for ch in str(cand).lower()))
            if k and len(k) > 2:
                out.setdefault(k, entry)
    return out

# Kirill -> lotin (INN/savdo nomini solishtirish uchun)
_CYR = {
    'а': 'a', 'б': 'b', 'в': 'v', 'г': 'g', 'д': 'd', 'е': 'e', 'ё': 'yo', 'ж': 'zh',
    'з': 'z', 'и': 'i', 'й': 'y', 'к': 'k', 'л': 'l', 'м': 'm', 'н': 'n', 'о': 'o',
    'п': 'p', 'р': 'r', 'с': 's', 'т': 't', 'у': 'u', 'ф': 'f', 'х': 'x', 'ц': 'ts',
    'ч': 'ch', 'ш': 'sh', 'щ': 'sch', 'ъ': '', 'ы': 'i', 'ь': '', 'э': 'e', 'ю': 'yu',
    'я': 'ya', 'ў': 'o', 'қ': 'q', 'ғ': 'g', 'ҳ': 'h',
}


def translit(s):
    if not s:
        return ''
    s = str(s).lower()
    return ''.join(_CYR.get(ch, ch) for ch in s)


def norm(s):
    """faqat harf/raqam, lotin."""
    return re.sub(r'[^a-z0-9]', '', translit(s))


_STRENGTH = re.compile(r'(\d+[.,]?\d*)\s*(мг/мл|мкг/доза|мг|мкг|г|мл|ме|ед|%|iu)', re.I)
_UNIT = {'мг': 'mg', 'мкг': 'mcg', 'г': 'g', 'мл': 'ml', 'ме': 'iu', 'ед': 'iu',
         '%': '%', 'мг/мл': 'mg/ml', 'мкг/доза': 'mcg/dose', 'iu': 'iu'}


def strengths(form):
    out = set()
    if not form:
        return out
    for num, unit in _STRENGTH.findall(str(form)):
        num = num.replace(',', '.').rstrip('0').rstrip('.') if '.' in num else num
        out.add(f"{num}{_UNIT.get(unit.lower(), unit.lower())}")
    return out


def main():
    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    inns = {}   # inn_lat -> {cyr, atc, count, strengths:set, trades:set, rx, foreign}
    trades = {}  # trade_lat -> inn_lat
    for i, r in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        trade = r[1]
        inn = r[5]
        atc = (r[6] or '').strip()
        form = r[9]
        rx = r[11]
        country = r[4] or ''
        substance = r[10]
        # INN yo'q bo'lsa -> ta'sir moddasi (col10, ' - ' dan oldingi qismi) yoki savdo nomi
        raw_inn = inn if inn and 'Отсутств' not in str(inn) else None
        subst = re.split(r'\s-\s|;', str(substance))[0].strip() if substance else ''
        name = raw_inn or subst or (str(trade).strip() if trade else '')
        if not name or 'отсутств' in name.lower():
            continue
        key = norm(name)
        if not key:
            continue
        e = inns.setdefault(key, {"cyr": name.strip()[:50], "atc": atc,
                                  "count": 0, "strengths": set(), "trades": set(),
                                  "rx": "", "foreign": True})
        e["count"] += 1
        e["strengths"] |= strengths(form)
        if trade:
            e["trades"].add(str(trade).strip().strip(',')[:40])
            trades[norm(trade)] = key
        if atc and not e["atc"]:
            e["atc"] = atc
        if rx and not e["rx"]:
            e["rx"] = "retseptsiz" if 'Без' in str(rx) else "retsept"
        if 'збекистан' in str(country) or 'zbekiston' in str(country):
            e["foreign"] = False  # mahalliy ishlab chiqarilgan ham bor

    # ATC indeks (birinchi 5 belgi = kimyoviy kichik guruh) -> muqobil
    atc_index = {}
    for key, e in inns.items():
        a5 = e["atc"][:5]
        if len(a5) >= 4:
            atc_index.setdefault(a5, []).append((key, e["count"]))
    for a5 in atc_index:
        atc_index[a5].sort(key=lambda x: -x[1])

    data = {
        "inns": {k: {"cyr": e["cyr"], "atc": e["atc"], "count": e["count"],
                     "strengths": sorted(e["strengths"]), "trades": sorted(e["trades"])[:5],
                     "rx": e["rx"], "local": not e["foreign"]}
                 for k, e in inns.items()},
        "trades": trades,
        "atc": {a: [k for k, _ in v[:8]] for a, v in atc_index.items()},
        "annulled": parse_annulled(),
    }
    print("  annullyatsiya:", len(data["annulled"]), "kalit")
    with open(OUT, "w") as f:
        json.dump(data, f, ensure_ascii=False)
    print(f"Yozildi: {OUT}")
    print(f"  noyob INN/modda: {len(inns)}, savdo nomlari: {len(trades)}, ATC guruh: {len(atc_index)}")
    return data


if __name__ == "__main__":
    data = main()
    # tez sinov: bizning dorilarni topa oladimi
    import difflib
    inns = data["inns"]; trades = data["trades"]
    keys = list(inns.keys())

    def find(q):
        k = norm(q)
        if k in inns:
            return k
        if k in trades:
            return trades[k]
        m = difflib.get_close_matches(k, keys, n=1, cutoff=0.82)
        if m:
            return m[0]
        for kk in keys:
            if kk.startswith(k) or (len(k) > 5 and k in kk):
                return kk
        return None

    print("\n--- SINOV ---")
    for q in ["Metformin", "Enalapril", "Amlodipin", "Atorvastatin", "Bisoprolol",
              "Omeprazol", "Metoprolol", "Furosemid", "Aspirin", "Insulin",
              "Levotiroksin", "Allopurinol", "Rozuvastatin", "Paratsetamol"]:
        k = find(q)
        if k:
            e = inns[k]
            print(f"  {q:14} -> ✓ {e['cyr'][:22]:22} | {e['count']:>3} ta | ATC {e['atc'][:7]:7} | doza: {', '.join(e['strengths'][:4])}")
        else:
            print(f"  {q:14} -> ✗ topilmadi")
